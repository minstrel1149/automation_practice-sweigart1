import openpyxl
import smtplib
import email
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
import pyinputplus as pyip

p1 = Path.cwd() / 'attachments'
p2 = Path.cwd() / 'result_attachments'

wb = openpyxl.load_workbook(p1 / 'duesRecords.xlsx')
sheet = wb.active
# max_column 메서드를 이용하여 마지막 열 지정
lastCol = sheet.max_column
# 최근 월 = sheet 내 1행 / 마지막 열의 값
lastMonth = sheet.cell(1, lastCol).value

# 미지불자의 이름과 이메일을 지정할 빈 딕셔너리 생성
unpaidMembers = {}
# 2행부터 마지막 행까지 순환문을 돌면서 payment 여부 확인
for row in range(2, sheet.max_row + 1):
    payment = sheet.cell(row, lastCol).value
    # 'paid'가 아닐 경우 이름(1열) 및 이메일(2열) 추가
    if payment != 'paid':
        name = sheet.cell(row, 1).value
        email = sheet.cell(row, 2).value
        unpaidMembers[name] = email

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
# pyinputplus를 사용하여 비밀번호 기입하도록 조치
smtpObj.login('kth1149@gmail.com', pyip.inputStr(prompt='Password: '))

# 딕셔너리의 items() 메서드를 순환시키면서 이름이랑 이메일 추출
for name, email in unpaidMembers.items():
    # MIMEMultipart() 활용하여 메시지 구성
    msg = MIMEMultipart()
    msg['To'] = email
    msg['From'] = 'tkim@iportfolio.co.kr'
    msg['cc'] = 'kth1149@gmail.com'
    msg['Subject'] = f'[IMPORTANT] {lastMonth} dues unpaid.'
    # html 스타일 활용하여 메일 본문 작성 -> HTML웹에디터 활용
    body = f'''
    <html>
    <p dir="ltr">{name} 님에게</p>
<p dir="ltr">&nbsp;</p>
<p dir="ltr">기록을 살펴보니 {name} 님께서 지난 달 회비를 지불하지 않으셨습니다. 확인하시고 조속히 회비를 지불해주실 것을 요청드립니다.</p>
<p dir="ltr">&nbsp;</p>
<p dir="ltr">관련하여 궁금하신 사항이 있으신 경우 언제든지 말씀 주세요.</p>
<p dir="ltr">&nbsp;</p>
<p dir="ltr">감사합니다. 좋은 하루 되세요 :)</p>
<p dir="ltr">김태화 드림</p>
<p dir="ltr">&nbsp;</p>
<p dir="ltr"><img class="CToWUd" src="https://ci5.googleusercontent.com/proxy/gBb4NHx_svsqyMi9VHqoMWbJ_7eo7qi2Sp0yaM9lRve2WklQGxhG5DkYtdD3RbAa8xJeJAjjZ-gbTsSfkGxCe-O1xfuoZjrWW3jjAV-R8l6xe2MqL9kWcCxlTudEuQCXlg=s0-d-e1-ft#https://ipf-release-dokdo.s3.ap-northeast-2.amazonaws.com/iPortfolio_96x40.png" data-bit="iit" /></p>
<p><span style="color: #000000;">김태화</span>| MP팀| 팀장 | 리드</p>
<p>Taehwa Kim | Management Planning Team | Team Leader | Lead Manager</p>
<p dir="ltr"><span style="color: #444444;">11F, 24 Namdaemun-ro 9-gil, Jung-gu, Seoul 04522, Korea</span></p>
<p dir="ltr">O. <span style="color: #444444;">+82-70-4880-3660</span> | M. <span style="color: #444444;">+82-10-4719-1149</span> | E. <a href="mailto:tkim@iportfolio.co.kr" target="_blank" rel="noopener"><span style="color: #444444;">tkim@iportfolio.co.kr</span></a> <a href="http://www.iportfolio.co.kr/" target="_blank" rel="noopener" data-saferedirecturl="https://www.google.com/url?q=http://www.iportfolio.co.kr/&amp;source=gmail&amp;ust=1659103992609000&amp;usg=AOvVaw1BunhL5jw5HED7sOq-ZY5h"><span style="color: #444444;">www.iportfolio.co.kr</span></a></p>
<p>&nbsp;</p>
<p dir="ltr">&nbsp;</p>
<p>The content of this email is confidential and intended for the recipient specified in message only. It is strictly forbidden to share any part of this message with any third party, without a written consent of the sender.</p>
</html>
    '''
    # msg에 MIMEText(body)를 추가 -> 'html' 형태이므로 두 번째 인자에 'html' 삽입
    msg.attach(MIMEText(body, 'html'))
    sendmailStatus = smtpObj.sendmail(msg['From'], msg['To'], msg.as_string())
    if sendmailStatus != {}:
        print('There was a problem.')

smtpObj.quit()