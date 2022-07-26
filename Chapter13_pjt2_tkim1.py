# pandas로도 쉽게 할 수 있을 것 같은데, 이번에는 openpyxl로
import openpyxl
from pathlib import Path

p1 = Path.cwd() / 'attachments'
p2 = Path.cwd() / 'result_attachments'

# 딕셔너리를 통해 농산물 업데이트 가격 설정
priceUpdates = {'Garlic':3.07, 'Celery':1.19, 'Lemon':1.27}

originWb = openpyxl.load_workbook(p1 / 'produceSales.xlsx')
sheet = originWb.active

# max_row 메서드를 range에 삽입하여 rowNum으로 순환
for rowNum in range(2, sheet.max_row + 1):
    # 먼저 A열에 있는 productName을 확인한 후
    productName = sheet.cell(rowNum, 1).value
    # productName이 priceUpdates 안에 들어가 있으면
    if productName in priceUpdates.keys():
        # B열에 있는 가격을 수정 -> priceUpdates[productName] 활용
        sheet.cell(rowNum, 2).value = priceUpdates[productName]

originWb.save(p2 / 'updatedProduceSales.xlsx')